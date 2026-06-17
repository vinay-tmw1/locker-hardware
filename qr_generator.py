"""
Generate a print-ready A4 PDF of QR code labels from the inventory Excel file.

Reads:
  - SKU_ID        (column A)
  - PRODUCT_NAME  (column B)
  - QR_BARCODE    (embedded images in column H, one per row)

Layout (per box):
  - Box size: 5.25cm wide x 2.1214cm tall
  - Grid: 4 columns x 14 rows per A4 page (edge-to-edge, no margins, no borders)
  - QR code on the left, 1mm padding, scaled to fill the box height
  - Product name (bold, max 8 chars + "...") and SKU_ID below it, on the right

Requirements:
  pip install openpyxl playwright --break-system-packages
  playwright install chromium

Usage:
  python generate_labels_pdf.py inventory_with_qr_FINAL_updated.xlsx output_labels.pdf
"""

import sys
import json
import base64
import html
import subprocess
import tempfile
import os

import openpyxl


ROWS_PER_PAGE = 14
COLS = 4
BOX_WIDTH_CM = 5.25
BOX_HEIGHT_CM = 2.1214
SHEET_HEIGHT_CM = ROWS_PER_PAGE * BOX_HEIGHT_CM  # ~29.70cm, fits A4 exactly


def extract_products(xlsx_path):
    """Read SKU_ID, PRODUCT_NAME, and embedded QR images from the workbook."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Map each embedded image to its 1-indexed row number
    row_to_img = {}
    for img in ws._images:
        excel_row = img.anchor._from.row + 1
        row_to_img[excel_row] = img._data()

    products = []
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if sku is None:
            continue
        img_data = row_to_img.get(r)
        if img_data is None:
            print(f"Warning: no QR image found for row {r} (SKU {sku}), skipping.")
            continue
        products.append({
            "sku": str(sku).strip(),
            "name": str(name).strip() if name else "",
            "img_b64": base64.b64encode(img_data).decode("ascii"),
        })
    return products


def truncate_name(name):
    """First 8 characters as-is, plus '...' if the name was longer."""
    if len(name) <= 8:
        return html.escape(name)
    return html.escape(name[:8]) + "..."


def build_html(products):
    cells_html = []
    for p in products:
        cell = f'''<div class="box">
  <img class="qr" src="data:image/png;base64,{p['img_b64']}" alt="QR"/>
  <div class="info">
    <div class="pname">{truncate_name(p['name'])}</div>
    <div class="psku">{html.escape(p['sku'])}</div>
  </div>
</div>'''
        cells_html.append(cell)

    per_page = ROWS_PER_PAGE * COLS
    pages = []
    for i in range(0, len(cells_html), per_page):
        chunk = cells_html[i:i + per_page]
        while len(chunk) < per_page:
            chunk.append('<div class="box empty"></div>')
        pages.append(chunk)

    page_blocks = [f'<div class="sheet">\n{"".join(chunk)}\n</div>' for chunk in pages]
    all_pages_html = "\n".join(page_blocks)

    css = f'''
@page {{ size: A4; margin: 0; }}
* {{ box-sizing: border-box; }}
html, body {{ margin: 0; padding: 0; }}
body {{ font-family: Arial, Helvetica, sans-serif; }}
.sheet {{
  width: 21cm;
  height: {SHEET_HEIGHT_CM}cm;
  display: grid;
  grid-template-columns: repeat({COLS}, {BOX_WIDTH_CM}cm);
  grid-template-rows: repeat({ROWS_PER_PAGE}, {BOX_HEIGHT_CM}cm);
  page-break-after: always;
}}
.sheet:last-child {{ page-break-after: auto; }}
.box {{
  border: none;
  box-sizing: border-box;
  width: {BOX_WIDTH_CM}cm;
  height: {BOX_HEIGHT_CM}cm;
  display: flex;
  align-items: center;
  overflow: hidden;
  padding: 0;
}}
.qr {{
  height: calc({BOX_HEIGHT_CM}cm - 2mm);
  width: calc({BOX_HEIGHT_CM}cm - 2mm);
  margin-left: 1mm;
  flex-shrink: 0;
  display: block;
  object-fit: contain;
}}
.info {{
  flex: 1;
  min-width: 0;
  padding-left: 2mm;
  padding-right: 1mm;
  display: flex;
  flex-direction: column;
  justify-content: center;
  overflow: hidden;
}}
.pname {{
  font-weight: bold;
  font-size: 9pt;
  line-height: 1.15;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
  text-align: left;
}}
.psku {{
  font-weight: normal;
  font-size: 7.5pt;
  line-height: 1.15;
  white-space: nowrap;
  overflow: hidden;
  text-align: left;
  margin-top: 0.8mm;
}}
'''

    return f'''<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Inventory QR Labels</title>
<style>{css}</style>
</head>
<body>
{all_pages_html}
</body>
</html>'''


def render_pdf(html_path, pdf_path):
    """Use Playwright (headless Chromium) to render the HTML to a pixel-accurate A4 PDF."""
    script = f'''
const {{ chromium }} = require('playwright');
(async () => {{
  const browser = await chromium.launch();
  const page = await browser.newPage();
  await page.goto('file://{html_path}');
  await page.pdf({{
    path: '{pdf_path}',
    format: 'A4',
    printBackground: true,
    margin: {{ top: '0', bottom: '0', left: '0', right: '0' }}
  }});
  await browser.close();
}})();
'''
    with tempfile.NamedTemporaryFile(mode="w", suffix=".js", delete=False) as f:
        f.write(script)
        script_path = f.name
    try:
        subprocess.run(["node", script_path], check=True)
    finally:
        os.unlink(script_path)


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_labels_pdf.py <input.xlsx> <output.pdf>")
        sys.exit(1)

    xlsx_path = os.path.abspath(sys.argv[1])
    pdf_path = os.path.abspath(sys.argv[2])
    html_path = os.path.splitext(pdf_path)[0] + ".html"

    products = extract_products(xlsx_path)
    print(f"Loaded {len(products)} products with QR images.")

    html_doc = build_html(products)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_doc)

    render_pdf(html_path, pdf_path)
    print(f"PDF saved to: {pdf_path}")
    print(f"HTML saved to: {html_path}")


if __name__ == "__main__":
    main()